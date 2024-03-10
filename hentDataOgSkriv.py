from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from openpyxl import Workbook, load_workbook

filnavn = 'testWebScraping.xlsx'
wb = load_workbook(filnavn) #workbook
ws = wb.active #worksheet

# Digg!! Fiksa problemet: https://www.google.com/url?sa=t&rct=j&q=&esrc=s&source=web&cd=&ved=2ahUKEwjNqefY8N2EAxXNGxAIHQgJAGAQwqsBegQICBAG&url=https%3A%2F%2Fwww.youtube.com%2Fwatch%3Fv%3DijT2sLVdnPM&usg=AOvVaw1AyCcQjN8u25YrCT605o7H&opi=89978449
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options,service=Service(ChromeDriverManager().install()))
# forhindrer at chrome lukker seg med en gang


url = 'https://www.proff.no/regnskap/searis-as/trondheim/faglig-vitenskapelig-og-teknisk-virksomhet/IGIP94N10N7'
driver.get(url)
# setter en ventetid på 10 sekunder som Selenium venter før det gir opp å finne et element som ikke er umiddelbart tilgjengelig.
# Uten dette kan koden feile hvis den prøver å finne elementer som ikke har lastet inn ennå.
driver.implicitly_wait(10)

time.sleep(1)  # Venter noen sekunder for å være sikker på at siden lastes inn og cookie-meldingen vises

# Deretter venter vi på at "Enig"-knappen blir tilgjengelig og klikker den
try:
    # Øk ventetiden i WebDriverWait hvis det er nødvendig
    enig_knapp = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.css-1hv8ibq'))
    )
    enig_knapp.click()
    print("Cookie-varsel er håndtert.")
except Exception as e:
    print(f"Kunne ikke finne 'Enig'-knappen: {e}")



def legg_til_tall(tall, lager):
    tall_u_mellomrom = tall.text.replace(' ', '')
    lager.append(tall_u_mellomrom)

def hent_data_resultat(nøkkeltall):
    table_container = driver.find_element(By.CLASS_NAME, 'MuiTableContainer-root')
    rader = table_container.find_elements(By.TAG_NAME, 'tr')
    data_lager = []
    for rad in rader:
        if nøkkeltall in rad.text:
            data = rad.find_elements(By.TAG_NAME, 'td')
            for tall in data:
                legg_til_tall(tall, data_lager)
            knapp_xpath = "//button[contains(@class,'MuiIconButton-root') and @aria-label='Previous years']"
            while True:
                try:
                    knapp = driver.find_element(By.XPATH, knapp_xpath)
                    if 'Mui-disabled' in knapp.get_attribute('class'):
                        print("Knappen er ikke trykkbar, går tilbake.")
                        fram_knapp_xpath = "//button[contains(@class,'MuiIconButton-root') and @aria-label='Latest years']"
                        while True:
                            try:
                                fram_knapp = driver.find_element(By.XPATH, fram_knapp_xpath)
                                if 'Mui-disabled' in fram_knapp.get_attribute('class'):
                                    print("Knappen er ikke trykkbar, avslutter løkken.")
                                    break
                                fram_knapp.click()
                                time.sleep(1)  # Vent litt for at siden skal laste etter klikk
                            except Exception as e:
                                print(f"Feil ved klikking: {e}")
                                break
                        break
                    knapp.click()
                    time.sleep(1)  # Vent litt for at siden skal laste etter klikk
                    nytt_tall = rad.find_elements(By.TAG_NAME, 'td')[4] #!!!IKKE HER!!!
                    legg_til_tall(nytt_tall, data_lager)
                except Exception as e:
                    print(f"Feil ved klikking: {e}")
                    break
    oppdatert_data_lager = []
    for tall in data_lager:
        if not tall.strip() == "": 
            nytt_tall = tall.replace('-', '0').replace('−', '-')
            oppdatert_data_lager.append(nytt_tall)  
    data_lager = oppdatert_data_lager
    return data_lager

def hent_data_balanse(nøkkeltall):
    table_container = driver.find_element(By.CLASS_NAME, 'MuiTableContainer-root')
    rader = table_container.find_elements(By.TAG_NAME, 'tr')
    data_lager = []
    for rad in rader:
        if nøkkeltall in rad.text:
            data = rad.find_elements(By.TAG_NAME, 'td')
            for tall in data:
                legg_til_tall(tall, data_lager)
            knapp_xpath = '//*[@id="scrollable-auto-tabpanel-1"]/div/div[1]/div/div/div[4]/div[3]/table/thead/tr/th[7]/button'
            while True:
                try:
                    knapp = driver.find_element(By.XPATH, knapp_xpath)
                    if 'Mui-disabled' in knapp.get_attribute('class'):
                        print("Knappen er ikke trykkbar, går tilbake.")
                        fram_knapp_xpath = '//*[@id="scrollable-auto-tabpanel-1"]/div/div[1]/div/div/div[4]/div[3]/table/thead/tr/th[2]/button'
                        while True:
                            try:
                                fram_knapp = driver.find_element(By.XPATH, fram_knapp_xpath)
                                if 'Mui-disabled' in fram_knapp.get_attribute('class'):
                                    print("Knappen er ikke trykkbar, avslutter løkken.")
                                    break
                                fram_knapp.click()
                                time.sleep(1)  # Vent litt for at siden skal laste etter klikk
                            except Exception as e:
                                print(f"Feil ved klikking: {e}")
                                break
                        break
                    knapp.click()
                    time.sleep(1)  # Vent litt for at siden skal laste etter klikk
                    nytt_tall = rad.find_elements(By.TAG_NAME, 'td')[4] #!!!IKKE HER!!!
                    legg_til_tall(nytt_tall, data_lager)
                except Exception as e:
                    print(f"Feil ved klikking: {e}")
                    break
    oppdatert_data_lager = []
    for tall in data_lager:
        if not tall.strip() == "": 
            nytt_tall = tall.replace('-', '0').replace('−', '-')
            oppdatert_data_lager.append(nytt_tall)  
    data_lager = oppdatert_data_lager
    return data_lager


a,b,c,d,e,f,g = 'Sum driftsinntekter', 'Sum salgsinntekter', 'Ordinære avskrivninger', 'Nedskrivning', 'Driftsresultat', 'Sum investeringer', 'Sum egenkapital'

driftsinntekter = hent_data_resultat(a)
salgsinntekter = hent_data_resultat(b)
avskrivninger = hent_data_resultat(c)
nedskrivninger = hent_data_resultat(d)
driftsresultat = hent_data_resultat(e)
EBITDA = []
for i in range(len(driftsresultat)):
    driftsresultat_i = driftsresultat[i]
    avskrivninger_i = avskrivninger[i]
    nedskrivninger_i = nedskrivninger[i]
    tall = int(driftsresultat_i) + int(avskrivninger_i) + int(nedskrivninger_i)
    EBITDA.append(str(tall))
investeringer = hent_data_balanse(g)
egenkapital = hent_data_balanse(f)
    
print(driftsinntekter)
print(salgsinntekter)
print(avskrivninger)
print(nedskrivninger)
print(driftsresultat)
print(EBITDA)
print(investeringer)
print(egenkapital)

driver.quit()


def skriv_til_excel(lager):
    rad = 110   
    for indeks, tall in enumerate(lager):
        kolonnebokstav = chr(78 - indeks)  # 66 er B i ASCII-tabellen
        ws[f'{kolonnebokstav}{rad}'] = tall  # Skriver til excel
        rad += 1
    wb.save(filnavn)

# skriv_til_excel(driftsinntekter)
# skriv_til_excel(salgsinntekter)
# skriv_til_excel(EBITDA)

