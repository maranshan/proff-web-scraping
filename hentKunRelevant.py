from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
import os

url = input("Lim inn URL: ")

filnavn = 'testWebScraping.xlsx'
wb = load_workbook(filnavn) #workbook
ws = wb.active #worksheet

# Digg!! Fiksa problemet: https://www.google.com/url?sa=t&rct=j&q=&esrc=s&source=web&cd=&ved=2ahUKEwjNqefY8N2EAxXNGxAIHQgJAGAQwqsBegQICBAG&url=https%3A%2F%2Fwww.youtube.com%2Fwatch%3Fv%3DijT2sLVdnPM&usg=AOvVaw1AyCcQjN8u25YrCT605o7H&opi=89978449
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options,service=Service(ChromeDriverManager().install()))
# forhindrer at chrome lukker seg med en gang


driver.get(url)
# setter en ventetid på 10 sekunder som Selenium venter før det gir opp å finne et element som ikke er umiddelbart tilgjengelig.
# Uten dette kan koden feile hvis den prøver å finne elementer som ikke har lastet inn ennå.
driver.implicitly_wait(10)

time.sleep(1)  # Venter noen sekunder for å være sikker på at siden lastes inn og cookie-meldingen vises

# Deretter venter vi på at "Enig"-knappen blir tilgjengelig og klikker den
try:
    # Øk ventetiden i WebDriverWait hvis det er nødvendig
    enig_knapp = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'button.css-1hv8ibq'))
    )
    enig_knapp.click()
    print("Cookie-varsel er håndtert.")
except Exception as e:
    print(f"Kunne ikke finne 'Enig'-knappen: {e}")

def legg_til_tall(tall, lager, nøkkel):
    tall_u_mellomrom = tall.text.replace(' ', '').replace('\n', '') 
    lager[nøkkel].append(tall_u_mellomrom)


def hent_data():
    nøkkeltall_liste = ['Sum driftsinntekter', 'Sum salgsinntekter', 'Ordinære avskrivninger', 'Nedskrivning', 'Driftsresultat', 'Sum investeringer', 'Sum egenkapital']
    table_container = driver.find_element(By.CLASS_NAME, 'MuiTableContainer-root')
    rader = table_container.find_elements(By.TAG_NAME, 'tr')
    data_lager = {nøkkel: [] for nøkkel in nøkkeltall_liste}
    for rad in rader:
        rad_text = rad.text
        for nøkkeltall in nøkkeltall_liste:
            if nøkkeltall in rad_text:
                data = rad.find_elements(By.TAG_NAME, 'td')
                for tall in data:
                    legg_til_tall(tall, data_lager, nøkkeltall)
    knapp_xpath = "//button[contains(@class,'MuiIconButton-root') and @aria-label='Previous years']"
    while True:
        try:
            knapp = driver.find_element(By.XPATH, knapp_xpath)
            if 'Mui-disabled' in knapp.get_attribute('class'):
                print('---> X')
                break
            knapp.click()
            time.sleep(1)  
            table_container = driver.find_element(By.CLASS_NAME, 'MuiTableContainer-root')
            rader = table_container.find_elements(By.TAG_NAME, 'tr')
            for rad in rader:
                rad_text = rad.text
                for nøkkeltall in nøkkeltall_liste:
                    if nøkkeltall in rad_text:
                        tall = rad.find_elements(By.TAG_NAME, 'td')[4]  
                        legg_til_tall(tall, data_lager, nøkkeltall)
        except Exception as e:
            print(f"Error on clicking: {e}")
            break
    for nøkkel, tall_liste in data_lager.items():
        oppdatert_tall_liste = []
        for tall in tall_liste:
            if tall.strip() != "":
                nytt_tall = tall.replace('-', '0').replace('−', '-')
                oppdatert_tall_liste.append(nytt_tall)
        data_lager[nøkkel] = oppdatert_tall_liste
    return data_lager

data = hent_data()
driver.quit()

a,b,c,d,e,f,g = 'Sum driftsinntekter', 'Sum salgsinntekter', 'Ordinære avskrivninger', 'Nedskrivning', 'Driftsresultat', 'Sum investeringer', 'Sum egenkapital'

driftsinntekter = data[a]
salgsinntekter = data[b]
avskrivninger = data[c]
nedskrivninger = data[d]
driftsresultat = data[e]
EBITDA = []
for i in range(len(driftsresultat)):
    driftsresultat_i = driftsresultat[i]
    avskrivninger_i = avskrivninger[i]
    nedskrivninger_i = nedskrivninger[i]
    tall = int(driftsresultat_i) + int(avskrivninger_i) + int(nedskrivninger_i)
    EBITDA.append(str(tall))
investeringer = data[f]
egenkapital = data[g]

rad = 1   
def skriv_til_excel(lager):
    global rad  
    for indeks, tall in enumerate(lager):
        kolonnebokstav = chr(75 - indeks)  
        ws[f'{kolonnebokstav}{rad}'] = tall  # Skriver til excel
    wb.save(filnavn)
    rad += 1

del wb[ws.title]

# Oppretter et nytt arbeidsark og setter det som aktivt
ws = wb.create_sheet("Sheet1")
wb.active = ws

skriv_til_excel(driftsinntekter)
skriv_til_excel(salgsinntekter)
skriv_til_excel(EBITDA)
skriv_til_excel(investeringer)
skriv_til_excel(egenkapital)

#Åpner automatisk opp excel fila
excel_fil_path = os.path.abspath(filnavn)
os.system(f"open '{excel_fil_path}'")

